from datetime import datetime
from os import path

import openpyxl

from src import dto, shared
from src.banks import interface


_months_map = {
    "января": "January",
    "февраля": "February",
    "марта": "March",
    "апреля": "April",
    "мая": "May",
    "июня": "June",
    "июля": "July",
    "августа": "August",
    "сентября": "September",
    "октября": "October",
    "ноября": "November",
    "декабря": "December",
}


def _deserialize_date(date: str) -> datetime:
    for rus_month, eng_month in _months_map.items():
        if rus_month in date:
            date = date.replace(rus_month, eng_month)

    return datetime.strptime(date, "%d %B %Y, %H:%M")


def _get_file_path(dir_path: str, file_name: str | None) -> str:
    if not file_name:
        file_name = shared.find_latest_file_name(dir_path)

    return path.join(dir_path, file_name)


_income_filter_words = [
    "Tochka Card2Card",
    "Тинькофф Банк",
    "Т-Банк (Тинькофф)",
]


def _filter_income(description: str) -> bool:
    for word in _income_filter_words:
        if word in description:
            return True

    return False


_outcome_filter_words = [
    "Дмитрий Сергеевич Г.",
    "Дмитрий Сергеевич Г",
]


def _filter_outcome(description: str) -> bool:
    for word in _outcome_filter_words:
        if word in description:
            return True

    return False


class Sber(interface.IBank):
    def __init__(
        self,
        income_dir_path: str,
        income_file_name: str | None,
        outcome_dir_path: str,
        outcome_file_name: str | None,
    ):
        self.income_file_path = _get_file_path(income_dir_path, income_file_name)
        self.outcome_file_path = _get_file_path(outcome_dir_path, outcome_file_name)

    async def get_name(self) -> str:
        return "SberBank"

    async def _get_income_operations(self) -> list[dto.IncomeOperation]:
        workbook = openpyxl.load_workbook(self.income_file_path)
        sheet = workbook.active
        income_operations = []

        for row in sheet.iter_rows(min_row=2):
            date = _deserialize_date(row[1].value)
            amount = row[6].value
            description = row[7].value
            if not (date or amount or description):
                break
            if _filter_income(description):
                continue

            income_operations.append(
                dto.IncomeOperation(
                    date=date,
                    amount=amount,
                    description=description,
                )
            )

        return income_operations

    async def _get_outcome_operations(self) -> list[dto.OutcomeOperation]:
        workbook = openpyxl.load_workbook(self.outcome_file_path)
        sheet = workbook.active
        outcome_operations = []

        for row in sheet.iter_rows(min_row=2):
            date = _deserialize_date(row[1].value)
            amount = row[6].value
            description = row[7].value
            if not (date or amount or description):
                break
            if _filter_outcome(description):
                continue

            outcome_operations.append(
                dto.OutcomeOperation(
                    date=date,
                    amount=amount,
                    description=description,
                )
            )

        return outcome_operations

    async def get_operations(self) -> dto.Operations:
        return dto.Operations(
            incomes=await self._get_income_operations(),
            outcomes=await self._get_outcome_operations(),
        )
